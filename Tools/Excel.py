import openpyxl, os

def writeExcel(path:str, data:dict, overwrite:bool=False):
    # Check path have file exist
    if not path or not os.path.exists(path) or overwrite:
        # Create a new workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write keys in the first row as headers
        for index, key in enumerate(data.keys(), start=1):
            sheet.cell(row=1, column=index, value=key)
    else:
        # Open existing workbook
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        # Check if keys in data match with headers in the file
        headers = [cell.value for cell in sheet[1]]
        if not all(key in headers for key in data.keys()):
            return "Error: Data keys do not match with file headers"

    # Find the last row in the sheet
    last_row = sheet.max_row

    # Write data in sheet
    for index, (key, value) in enumerate(data.items(), start=1):
        sheet.cell(row=last_row + 1, column=index, value=value)

    # Save file
    workbook.save(path if path else 'New Dataset.xlsx')
    return "Save data in file success"